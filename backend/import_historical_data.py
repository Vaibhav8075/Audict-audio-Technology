import sys
import os
import datetime
from sqlalchemy.orm import Session
import openpyxl

# Add backend directory to path so we can import local modules
backend_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(backend_dir)

# Override DATABASE_URL to use the backend directory explicitly
os.environ["DATABASE_URL"] = f"sqlite:///{os.path.join(backend_dir, 'dcm_audit.db')}"

from database.connection import SessionLocal, engine
from models import (
    Base, User, Audit, Recording, AIAnalysis, FeedbackForm, 
    FeedbackQuestion, FeedbackAnswer, FeedbackQuestionAnswer, 
    UserRole, AuditStatus, FeedbackQuestionType, SentimentLabel
)
from auth_utils import get_password_hash

# Ensure tables exist
Base.metadata.create_all(bind=engine)

def parse_call_time(val):
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
            try:
                return datetime.datetime.strptime(val.strip(), fmt)
            except ValueError:
                pass
    return datetime.datetime.utcnow()

def parse_duration_to_seconds(val):
    if isinstance(val, datetime.time):
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        try:
            parts = list(map(int, val.strip().split(':')))
            if len(parts) == 3:
                return parts[0] * 3600 + parts[1] * 60 + parts[2]
            elif len(parts) == 2:
                return parts[0] * 60 + parts[1]
        except Exception:
            pass
    return 0

def import_data():
    excel_path = r"c:\Users\vibhu\Downloads\Weekly_Call_Audit_Recording_Sheet_1206.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return

    print("Loading workbook...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    db: Session = SessionLocal()

    # Pre-hash a default password to speed up importing
    print("Generating default password hash...")
    default_hashed_pwd = get_password_hash("password123")

    # Keep a local cache of created users to avoid DB hits
    users_cache = {}
    
    # Pre-load existing users from DB into cache
    for u in db.query(User).all():
        users_cache[u.email] = u.id

    def get_or_create_user(name, role):
        if not name:
            name = "unknown"
        name_clean = str(name).strip()
        email = f"{name_clean.lower()}@dcm.local"
        if email in users_cache:
            return users_cache[email]
        
        # Create user
        user = User(
            email=email,
            full_name=name_clean.capitalize(),
            hashed_password=default_hashed_pwd,
            role=role,
            department="Operations",
            is_active=True
        )
        db.add(user)
        db.flush()
        users_cache[email] = user.id
        return user.id

    # Create the Default Feedback Form if not exists
    form = db.query(FeedbackForm).filter(FeedbackForm.title == "Call Audit Assessment Form").first()
    if not form:
        print("Creating Call Audit Assessment Form...")
        admin_id = get_or_create_user("system", UserRole.admin)
        form = FeedbackForm(
            title="Call Audit Assessment Form",
            description="Standard Quality Assurance assessment form matching the Excel audit sheet.",
            created_by=admin_id
        )
        db.add(form)
        db.flush()
        
        # Define questions matching the sheet columns
        questions_meta = [
            ("Welcoming Customer", FeedbackQuestionType.rating),
            ("Confirmed Caller Details", FeedbackQuestionType.rating),
            ("Listening to Customer", FeedbackQuestionType.rating),
            ("Tone & Empathy", FeedbackQuestionType.rating),
            ("Clear Communication", FeedbackQuestionType.rating),
            ("Service Delivery / Promptness", FeedbackQuestionType.rating),
            ("Delivery of Right Information", FeedbackQuestionType.rating),
            ("Genuine Ticket Raising", FeedbackQuestionType.yes_no),
            ("Able to achieve the outcome", FeedbackQuestionType.yes_no),
        ]
        
        for idx, (q_text, q_type) in enumerate(questions_meta):
            q = FeedbackQuestion(
                form_id=form.id,
                question_text=q_text,
                question_type=q_type,
                order_index=idx
            )
            db.add(q)
        db.flush()
        print(f"Created Form with ID: {form.id}")

    # Load questions to reference their IDs
    questions = db.query(FeedbackQuestion).filter(FeedbackQuestion.form_id == form.id).order_by(FeedbackQuestion.order_index).all()

    total_audits_imported = 0

    print("Scanning sheets...")
    for sheet_name in wb.sheetnames:
        # We process sheets containing historical week/work audit reports
        if not (sheet_name.startswith("Working-") or sheet_name.startswith("Week") or "Working" in sheet_name or sheet_name.isdigit() or sheet_name == "2603_0104"):
            continue
            
        sheet = wb[sheet_name]
        print(f"\nProcessing Sheet: {sheet_name}")
        
        # Find the header row (Sno, Agent ID, Phone No)
        header_row_idx = None
        rows = list(sheet.iter_rows(values_only=True))
        
        for r_idx, row in enumerate(rows[:10]):
            row_str = [str(x).lower() if x is not None else "" for x in row]
            if "sno" in row_str and ("agent id" in row_str or "agent" in row_str) and ("phone" in row_str or "phone no" in row_str):
                header_row_idx = r_idx
                break
                
        if header_row_idx is None:
            print(f"Skipping {sheet_name} (No header row found)")
            continue
            
        header = [str(x).strip() if x is not None else "" for x in rows[header_row_idx]]
        print(f"Header row found at index {header_row_idx + 1}")
        
        # Map header columns to indexes
        col_mapping = {}
        for idx, col_name in enumerate(header):
            col_name_lower = col_name.lower().replace("\n", " ")
            if "sno" in col_name_lower:
                col_mapping["sno"] = idx
            elif "weeknum" in col_name_lower or "week" in col_name_lower:
                col_mapping["week"] = idx
            elif "agent id" in col_name_lower or "agent" in col_name_lower:
                col_mapping["agent"] = idx
            elif "phone" in col_name_lower:
                col_mapping["phone"] = idx
            elif "process" in col_name_lower:
                col_mapping["process"] = idx
            elif "call type" in col_name_lower:
                col_mapping["call_type"] = idx
            elif "call time" in col_name_lower or "date" in col_name_lower:
                col_mapping["call_time"] = idx
            elif "duration" in col_name_lower:
                col_mapping["duration"] = idx
            elif "assigned" in col_name_lower or "auditor" in col_name_lower:
                col_mapping["auditor"] = idx
            elif "file" in col_name_lower or "path" in col_name_lower or "wav" in col_name_lower:
                col_mapping["file_path"] = idx
            elif "marks obtained" in col_name_lower:
                col_mapping["marks_obtained"] = idx
            elif "%" in col_name_lower or "percentage" in col_name_lower:
                col_mapping["percentage"] = idx
                
        # Also need specific index for ratings questions (Welcoming, etc.)
        # Let's search by string match in the header
        eval_cols = [
            ("Welcoming Customer", "welcoming"),
            ("Confirmed Caller Details", "confirmed caller"),
            ("Listening to Customer", "listening"),
            ("Tone & Empathy", "tone"),
            ("Clear Communication", "clear comm"),
            ("Service Delivery / Promptness", "promptness"),
            ("Delivery of Right Information", "delivery of right"),
            ("Genuine Ticket Raising", "ticket raising"),
            ("Able to achieve the outcome", "achieve the outcome"),
        ]
        
        eval_indices = []
        for q_name, search_key in eval_cols:
            col_idx = None
            for idx, col_name in enumerate(header):
                if search_key in col_name.lower():
                    # Pick the qualitative column (typically comes first before the marks columns)
                    col_idx = idx
                    break
            eval_indices.append(col_idx)

        # Print identified columns
        print(f"Column Mapping: {col_mapping}")
        print(f"Eval Columns Indices: {eval_indices}")
        
        # Read audit rows
        sheet_audits_count = 0
        for r_idx in range(header_row_idx + 1, len(rows)):
            row = rows[r_idx]
            
            # Check if row is empty or sno is null
            sno_val = row[col_mapping.get("sno", 0)] if "sno" in col_mapping else None
            if sno_val is None or str(sno_val).strip() == "":
                # End of sheet table
                break
                
            try:
                sno = int(sno_val)
            except ValueError:
                # Header group or note row
                continue
                
            # Extract fields
            agent_name = row[col_mapping["agent"]] if "agent" in col_mapping else "agent"
            phone_val = str(row[col_mapping["phone"]]).strip() if "phone" in col_mapping else "0000000000"
            process = str(row[col_mapping["process"]]).strip() if "process" in col_mapping else "Campaign"
            call_type = str(row[col_mapping["call_type"]]).strip() if "call_type" in col_mapping else "OUTGOING"
            call_time_val = row[col_mapping["call_time"]] if "call_time" in col_mapping else None
            duration_val = row[col_mapping["duration"]] if "duration" in col_mapping else None
            auditor_name = str(row[col_mapping["auditor"]]).strip() if "auditor" in col_mapping else "Auditor"
            file_path_val = str(row[col_mapping["file_path"]]).strip() if "file_path" in col_mapping else ""
            marks_val = row[col_mapping["marks_obtained"]] if "marks_obtained" in col_mapping else 0
            pct_val = row[col_mapping["percentage"]] if "percentage" in col_mapping else 0
            
            # Date & Duration
            call_date = parse_call_time(call_time_val)
            duration_sec = parse_duration_to_seconds(duration_val)
            
            # Generate a unique audit_id
            week_num = row[col_mapping["week"]] if "week" in col_mapping else "51"
            clean_sheet_name = "".join(c for c in sheet_name if c.isalnum())
            audit_id_str = f"AUD-{clean_sheet_name}-{sno}"
            
            # Check duplicate in DB
            exists = db.query(Audit).filter(Audit.audit_id == audit_id_str).first()
            if exists:
                continue
                
            # Get or create users
            agent_user_id = get_or_create_user(agent_name, UserRole.employee)
            auditor_user_id = get_or_create_user(auditor_name, UserRole.admin)
            
            # Create Audit
            audit = Audit(
                audit_id=audit_id_str,
                client_name=process,
                employee_id=agent_user_id,
                call_date=call_date,
                call_duration=duration_sec,
                status=AuditStatus.completed,
                notes=f"Imported from {sheet_name} SNo {sno}."
            )
            db.add(audit)
            db.flush()
            
            # Create Recording record (marked expired)
            recording = Recording(
                audit_id=audit.id,
                file_name=os.path.basename(file_path_val) if file_path_val else f"call_{phone_val}.wav",
                file_path=None,
                file_size=0,
                duration=duration_sec,
                is_expired=True,
                expires_at=datetime.datetime.utcnow(),
                uploaded_by=auditor_user_id
            )
            db.add(recording)
            
            # Create AI Analysis summary matching the score
            try:
                score_pct = float(pct_val)
                # handle if sheet stored it as e.g. 0.6 instead of 60.0
                if score_pct <= 1.0:
                    score_pct = score_pct * 100
            except (ValueError, TypeError):
                score_pct = 0.0
                
            ai_analysis = AIAnalysis(
                audit_id=audit.id,
                transcription="[Transcripts not imported for historical data]",
                call_summary=f"Historical Call Audit score: {round(score_pct, 2)}%",
                call_quality_score=score_pct,
                customer_satisfaction_score=score_pct,
                sentiment_label=SentimentLabel.neutral if score_pct >= 50 else SentimentLabel.negative,
                model_used="Manual Import"
            )
            db.add(ai_analysis)
            
            # Create Feedback Answers (Manual scores)
            feedback = FeedbackAnswer(
                audit_id=audit.id,
                form_id=form.id,
                submitted_by_id=auditor_user_id,
                overall_rating=score_pct / 20.0 if score_pct else 0.0, # Convert to 0-5 rating
                comments=f"Score Obtained: {marks_val}. Audited by {auditor_name}."
            )
            db.add(feedback)
            db.flush()
            
            # Save question answers
            for q_idx, q_col_idx in enumerate(eval_indices):
                if q_col_idx is None or q_col_idx >= len(row):
                    continue
                ans_val = row[q_col_idx]
                if ans_val is not None:
                    qa = FeedbackQuestionAnswer(
                        submission_id=feedback.id,
                        question_id=questions[q_idx].id,
                        answer_value=str(ans_val).strip()
                    )
                    db.add(qa)
                    
            sheet_audits_count += 1
            total_audits_imported += 1
            
            if sheet_audits_count % 100 == 0:
                db.commit()
                print(f"  Imported {sheet_audits_count} rows...")
                
        db.commit()
        print(f"Completed sheet {sheet_name}. Imported {sheet_audits_count} audits.")

    print(f"\nSuccessfully imported a total of {total_audits_imported} audits into the database.")
    db.close()

if __name__ == "__main__":
    import_data()
